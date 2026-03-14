from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .parser import ParseResult

COLUMN_ORDER = [
    ("Sr No", None, None),
    ("College Code", None, None),
    ("College Name", None, None),
    ("City", None, None),
    ("District", None, None),
    ("College Type", None, None),
    ("Minority Status", None, None),
    ("Branch", None, None),
    ("GOPENS Rank", "GOPENS", "rank"),
    ("GOPENS %", "GOPENS", "pct"),
    ("LOPENS Rank", "LOPENS", "rank"),
    ("LOPENS %", "LOPENS", "pct"),
    ("GSCS Rank", "GSCS", "rank"),
    ("GSCS %", "GSCS", "pct"),
    ("GSTS Rank", "GSTS", "rank"),
    ("GSTS %", "GSTS", "pct"),
    ("GVJS Rank", "GVJS", "rank"),
    ("GVJS %", "GVJS", "pct"),
    ("GNT1S Rank", "GNT1S", "rank"),
    ("GNT1S %", "GNT1S", "pct"),
    ("GNT2S Rank", "GNT2S", "rank"),
    ("GNT2S %", "GNT2S", "pct"),
    ("GNT3S Rank", "GNT3S", "rank"),
    ("GNT3S %", "GNT3S", "pct"),
    ("GOBCS Rank", "GOBCS", "rank"),
    ("GOBCS %", "GOBCS", "pct"),
    ("GSEBCS Rank", "GSEBCS", "rank"),
    ("GSEBCS %", "GSEBCS", "pct"),
    ("LSCS Rank", "LSCS", "rank"),
    ("LSCS %", "LSCS", "pct"),
    ("LSTS Rank", "LSTS", "rank"),
    ("LSTS %", "LSTS", "pct"),
    ("LVJS Rank", "LVJS", "rank"),
    ("LVJS %", "LVJS", "pct"),
    ("LNT1S Rank", "LNT1S", "rank"),
    ("LNT1S %", "LNT1S", "pct"),
    ("LNT2S Rank", "LNT2S", "rank"),
    ("LNT2S %", "LNT2S", "pct"),
    ("LNT3S Rank", "LNT3S", "rank"),
    ("LNT3S %", "LNT3S", "pct"),
    ("LOBCS Rank", "LOBCS", "rank"),
    ("LOBCS %", "LOBCS", "pct"),
    ("LSEBCS Rank", "LSEBCS", "rank"),
    ("LSEBCS %", "LSEBCS", "pct"),
    ("TFWS Rank", "TFWS", "rank"),
    ("TFWS %", "TFWS", "pct"),
    ("EWS Rank", "EWS", "rank"),
    ("EWS %", "EWS", "pct"),
    ("PWDOPENS Rank", "PWDOPENS", "rank"),
    ("PWDOPENS %", "PWDOPENS", "pct"),
    ("PWDOBCS Rank", "PWDOBCS", "rank"),
    ("PWDOBCS %", "PWDOBCS", "pct"),
    ("PWDRSCS Rank", "PWDRSCS", "rank"),
    ("PWDRSCS %", "PWDRSCS", "pct"),
    ("DEFOPENS Rank", "DEFOPENS", "rank"),
    ("DEFOPENS %", "DEFOPENS", "pct"),
    ("DEFOBCS Rank", "DEFOBCS", "rank"),
    ("DEFOBCS %", "DEFOBCS", "pct"),
    ("DEFROBCS Rank", "DEFROBCS", "rank"),
    ("DEFROBCS %", "DEFROBCS", "pct"),
    ("ORPHAN Rank", "ORPHAN", "rank"),
    ("ORPHAN %", "ORPHAN", "pct"),
]

GROUP_FILLS = {
    "identity": "BDD7EE",
    "open": "C6EFCE",
    "reserved": "FCE4D6",
    "ladies": "E2EFDA",
    "special": "FFEB9C",
}

TEXT_COLUMNS = {"College Code"}
LEFT_ALIGN_COLUMNS = {"College Name", "City", "District", "College Type", "Minority Status", "Branch"}
HOW_TO_USE_LINES = [
    "HOW TO USE THIS FILE FOR STUDENT COUNSELING",
    "",
    "Step 1: Ask student for:",
    "  - Their MHT-CET Percentile",
    "  - Their Category (GOPENS / GOBCS / GSCS / GSTS / GNT1S / GNT2S / GNT3S / GSEBCS / EWS / TFWS)",
    "  - Preferred City or District",
    "  - Preferred Branch (CS / IT / ENTC etc.)",
    "  - Gender (if girl student, use LOPENS / ladies category columns)",
    "",
    'Step 2: In the "CAP Round I Cutoffs" sheet:',
    "  - Filter Column D (City) OR Column E (District) by student's preference",
    "  - Filter Column H (Branch) by student's preferred branch",
    "  - Filter Column F (College Type) if student wants only Government colleges",
    "",
    "Step 3: Sort student's category column (Percentile) in DESCENDING order",
    "",
    "Step 4: Find rows where student's percentile is GREATER THAN the cutoff percentile",
    "  These are the colleges the student CAN get admission in (safe options)",
    '  Colleges just below their percentile are "borderline" options',
    "",
    "Step 5: Consider College Type:",
    "  - Government colleges = low fees (~20,000/year)",
    "  - Government-Aided = medium fees",
    "  - Private (Unaided) = high fees (1.5-2 lakh/year)",
    "",
    "Category Column Reference:",
    "  GOPENS = General / Open category",
    "  GOBCS  = OBC (Other Backward Class)",
    "  GSCS   = SC (Scheduled Caste)",
    "  GSTS   = ST (Scheduled Tribe)",
    "  GVJS   = VJ / DT (Vimukta Jati / Denotified Tribe)",
    "  GNT1S  = NT1 (Nomadic Tribe 1 - Matang)",
    "  GNT2S  = NT2 (Nomadic Tribe 2)",
    "  GNT3S  = NT3 (Nomadic Tribe 3)",
    "  GSEBCS = SEBC (Maratha Reservation)",
    "  LOPENS = Ladies Open (girl students - General)",
    "  TFWS   = Tuition Fee Waiver (income below 8 lakh)",
    "  EWS    = Economically Weaker Section",
]


@dataclass(slots=True)
class WorkbookSummary:
    sheet_name: str
    output_filename: str
    row_count: int


def get_group(header: str) -> str:
    identity_columns = {
        "Sr No",
        "College Code",
        "College Name",
        "City",
        "District",
        "College Type",
        "Minority Status",
        "Branch",
    }
    if header in identity_columns:
        return "identity"
    if header.startswith("GOPENS") or header.startswith("LOPENS"):
        return "open"
    if any(header.startswith(prefix) for prefix in ["GSCS", "GSTS", "GVJS", "GNT", "GOBCS", "GSEBCS"]):
        return "reserved"
    if header.startswith("L") and not header.startswith("LOPENS"):
        return "ladies"
    return "special"


def _set_column_widths(worksheet: openpyxl.worksheet.worksheet.Worksheet) -> None:
    for index, (header, _, _) in enumerate(COLUMN_ORDER, start=1):
        if header == "Sr No":
            width = 6
        elif header == "College Code":
            width = 10
        elif header == "College Name":
            width = 45
        elif header == "City":
            width = 18
        elif header == "District":
            width = 18
        elif header == "College Type":
            width = 20
        elif header == "Minority Status":
            width = 22
        elif header == "Branch":
            width = 35
        elif header.endswith("Rank"):
            width = 10
        else:
            width = 12
        worksheet.column_dimensions[get_column_letter(index)].width = width


def _write_headers(worksheet: openpyxl.worksheet.worksheet.Worksheet) -> None:
    for column_index, (header, _, _) in enumerate(COLUMN_ORDER, start=1):
        cell = worksheet.cell(row=1, column=column_index, value=header)
        cell.fill = PatternFill("solid", fgColor=GROUP_FILLS[get_group(header)])
        cell.font = Font(bold=True, size=11)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    worksheet.row_dimensions[1].height = 40
    worksheet.freeze_panes = "A2"


def _write_data_rows(worksheet: openpyxl.worksheet.worksheet.Worksheet, parse_result: ParseResult) -> None:
    for row_index, row in enumerate(parse_result.rows, start=2):
        fill = PatternFill("solid", fgColor="FFFFFF" if row_index % 2 == 0 else "F7F7F7")
        values = [
            row_index - 1,
            row.college_code,
            row.college_name,
            row.city,
            row.district,
            row.college_type,
            row.minority_status,
            row.branch_name,
        ]
        for _, category_code, field in COLUMN_ORDER[8:]:
            values.append(row.data.get(category_code, {}).get(field) if category_code else None)

        for column_index, value in enumerate(values, start=1):
            header = COLUMN_ORDER[column_index - 1][0]
            cell = worksheet.cell(row=row_index, column=column_index, value=value)
            cell.fill = fill
            cell.alignment = Alignment(
                horizontal="left" if header in LEFT_ALIGN_COLUMNS else "center",
                vertical="center",
            )
            if header in TEXT_COLUMNS:
                cell.number_format = "@"
            elif header.endswith("Rank") and value is not None:
                cell.number_format = "0"
            elif header.endswith("%") and value is not None:
                cell.number_format = "0.0000000"
        worksheet.row_dimensions[row_index].height = 18


def _write_how_to_use_sheet(workbook: openpyxl.Workbook) -> None:
    worksheet = workbook.create_sheet("How to Use")
    worksheet["A1"] = "\n".join(HOW_TO_USE_LINES)
    worksheet["A1"].font = Font(size=11)
    worksheet["A1"].alignment = Alignment(vertical="top", wrap_text=True)
    worksheet["A1"].fill = PatternFill("solid", fgColor="FFF9E6")
    worksheet.column_dimensions["A"].width = 110
    worksheet.row_dimensions[1].height = 520


def write_excel(parse_result: ParseResult, output_path: str | Path) -> WorkbookSummary:
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = parse_result.sheet_name[:31]

    _write_headers(worksheet)
    _write_data_rows(worksheet, parse_result)
    _set_column_widths(worksheet)
    worksheet.auto_filter.ref = f"A1:{get_column_letter(len(COLUMN_ORDER))}{max(parse_result.rows_found, 1) + 1}"

    _write_how_to_use_sheet(workbook)
    workbook.save(output_path)

    return WorkbookSummary(
        sheet_name=worksheet.title,
        output_filename=output_path.name,
        row_count=parse_result.rows_found,
    )
