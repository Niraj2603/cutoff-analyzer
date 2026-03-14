from __future__ import annotations

from pathlib import Path
import shutil

import openpyxl

from backend.excel_writer import COLUMN_ORDER, write_excel
from backend.parser import ParseResult, ParsedRow


def build_parse_result() -> ParseResult:
    return ParseResult(
        round_label="CAP Round I",
        academic_year="2025-26",
        sheet_name="CAP Round I Cutoffs",
        output_filename="MHT_CET_CAP_Round_I_Cutoffs_2025-26.xlsx",
        rows=[
            ParsedRow(
                college_code="01002",
                college_name="Government College of Engineering, Amravati",
                city="Amravati",
                district="Amravati",
                college_type="Government",
                minority_status="General",
                home_university="Sant Gadge Baba Amravati University",
                branch_code="0100224210",
                branch_name="Computer Science and Engineering",
                data={
                    "GOPENS": {"rank": 9196, "pct": 97.3737374},
                    "GOBCS": {"rank": 16679, "pct": 95.21},
                    "TFWS": {"rank": 7401, "pct": 97.86},
                },
            )
        ],
    )


def make_work_dir(test_name: str) -> Path:
    work_dir = Path("backend") / "tmp" / "test-artifacts" / test_name
    shutil.rmtree(work_dir, ignore_errors=True)
    work_dir.mkdir(parents=True, exist_ok=True)
    return work_dir


def test_write_excel_creates_expected_headers_and_formats() -> None:
    output_path = make_work_dir("excel-headers") / "cutoffs.xlsx"
    parse_result = build_parse_result()

    summary = write_excel(parse_result, output_path)
    workbook = openpyxl.load_workbook(output_path)
    worksheet = workbook["CAP Round I Cutoffs"]

    assert summary.row_count == 1
    assert worksheet.max_column == len(COLUMN_ORDER)
    assert [worksheet.cell(row=1, column=index).value for index in range(1, len(COLUMN_ORDER) + 1)] == [
        column[0] for column in COLUMN_ORDER
    ]
    assert worksheet["A2"].value == 1
    assert worksheet["B2"].value == "01002"
    assert worksheet["B2"].number_format == "@"
    assert worksheet["J2"].number_format == "0.0000000"
    assert worksheet["AS2"].value == 7401
    assert worksheet["AC2"].value is None
    assert worksheet.freeze_panes == "A2"
    assert worksheet.row_dimensions[1].height == 40
    assert worksheet.row_dimensions[2].height == 18
    assert worksheet["A1"].fill.fgColor.rgb.endswith("BDD7EE")
    assert worksheet["I1"].fill.fgColor.rgb.endswith("C6EFCE")


def test_write_excel_creates_how_to_use_sheet() -> None:
    output_path = make_work_dir("excel-guide") / "guide.xlsx"
    write_excel(build_parse_result(), output_path)

    workbook = openpyxl.load_workbook(output_path)
    guide_sheet = workbook["How to Use"]
    assert "HOW TO USE THIS FILE FOR STUDENT COUNSELING" in guide_sheet["A1"].value
    assert "GOBCS  = OBC (Other Backward Class)" in guide_sheet["A1"].value
